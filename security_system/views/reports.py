"""
Reports view — generate and export reports (Excel, PDF).
"""
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from ..auth import firebase_auth_required, ROLE_ADMIN, ROLE_SUPERVISOR
from ..firebase import (
    get_logs_ref, get_personnel_ref, get_areas_ref, get_shifts_ref,
    query_to_list,
)


@firebase_auth_required(allowed_roles=[ROLE_ADMIN, ROLE_SUPERVISOR])
def reports_view(request):
    """Render the reports page with filters."""
    personnel_ref = get_personnel_ref()
    areas_ref = get_areas_ref()
    shifts_ref = get_shifts_ref()

    context = {
        'all_personnel': query_to_list(personnel_ref.get()),
        'all_areas': query_to_list(areas_ref.get()),
        'all_shifts': query_to_list(shifts_ref.get()),
    }

    return render(request, 'security/reports.html', context)


def _get_filtered_logs(request):
    """Helper to get filtered logs based on query params."""
    logs_ref = get_logs_ref()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    query = logs_ref.order_by('entryTime', direction='DESCENDING')

    if date_from:
        try:
            dt_from = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.where('entryTime', '>=', dt_from)
        except ValueError:
            pass

    if date_to:
        try:
            dt_to = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.where('entryTime', '<=', dt_to)
        except ValueError:
            pass

    all_logs = query_to_list(query.limit(5000).get())

    # Get reference data
    all_personnel = query_to_list(get_personnel_ref().get())
    all_areas = query_to_list(get_areas_ref().get())
    all_shifts = query_to_list(get_shifts_ref().get())

    personnel_map = {p['id']: p for p in all_personnel}
    areas_map = {a['id']: a for a in all_areas}
    shifts_map = {s['id']: s for s in all_shifts}

    # Enrich
    for log in all_logs:
        person = personnel_map.get(log.get('personnelId'), {})
        log['personnelName'] = f"{person.get('firstName', '?')} {person.get('lastName', '')}"
        log['personnelDni'] = person.get('documentId', '')
        log['cargo'] = person.get('cargo', '')

        area = areas_map.get(log.get('destinationAreaId'), {})
        log['areaName'] = area.get('name', '—')

        shift = shifts_map.get(log.get('shiftId'), {})
        log['shiftName'] = shift.get('name', '—')

        # Duration
        if log.get('exitTime') and log.get('entryTime'):
            entry = log['entryTime']
            exit_t = log['exitTime']
            if hasattr(entry, 'timestamp') and hasattr(exit_t, 'timestamp'):
                delta = exit_t - entry
                hours = delta.total_seconds() / 3600
                log['duration'] = f"{int(hours)}h {int((hours % 1) * 60)}m"
                log['durationHours'] = round(hours, 2)
            else:
                log['duration'] = '—'
                log['durationHours'] = 0
        else:
            log['duration'] = 'En curso'
            log['durationHours'] = 0

    return all_logs, date_from, date_to


@firebase_auth_required(allowed_roles=[ROLE_ADMIN, ROLE_SUPERVISOR])
def export_excel(request):
    """Export filtered logs to Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    logs, date_from, date_to = _get_filtered_logs(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Seguridad"

    # Header styling
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = 'REPORTE DE CONTROL DE ENTRADA Y SALIDA — SENATI SEGURIDAD'
    ws['A1'].font = Font(name="Calibri", size=14, bold=True, color="1B3A5C")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Período: {date_from or "Inicio"} — {date_to or "Actual"}'
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Guardia', 'DNI', 'Cargo', 'Turno',
        'Fecha/Hora Entrada', 'Fecha/Hora Salida',
        'Duración', 'Motivo Salida', 'Área Destino', 'Estado'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    for row_idx, log in enumerate(logs, 5):
        entry_time = log.get('entryTime', '')
        exit_time = log.get('exitTime', '')

        if hasattr(entry_time, 'strftime'):
            entry_time = entry_time.strftime('%d/%m/%Y %H:%M')
        if hasattr(exit_time, 'strftime'):
            exit_time = exit_time.strftime('%d/%m/%Y %H:%M')
        elif not exit_time:
            exit_time = '—'

        reason = log.get('exitReason', '—') or '—'
        if reason == 'Otro' and log.get('exitReasonDetail'):
            reason = f"Otro: {log['exitReasonDetail']}"

        values = [
            log.get('personnelName', ''),
            log.get('personnelDni', ''),
            log.get('cargo', ''),
            log.get('shiftName', ''),
            entry_time,
            exit_time,
            log.get('duration', ''),
            reason,
            log.get('areaName', ''),
            'Dentro' if log.get('status') == 'entered' else 'Salió',
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(value))
            cell.font = cell_font
            cell.border = thin_border

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value and not hasattr(cell, 'coordinate'): # simpler way to skip MergedCells
                pass
            elif cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

    # Add auto-filter
    ws.auto_filter.ref = f"A4:J{len(logs) + 4}"

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_seguridad_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


@firebase_auth_required(allowed_roles=[ROLE_ADMIN, ROLE_SUPERVISOR])
def export_pdf(request):
    """Export filtered logs to PDF."""
    logs, date_from, date_to = _get_filtered_logs(request)

    # Build HTML for the PDF
    rows_html = ""
    for log in logs:
        entry_time = log.get('entryTime', '')
        exit_time = log.get('exitTime', '')

        if hasattr(entry_time, 'strftime'):
            entry_time = entry_time.strftime('%d/%m/%Y %H:%M')
        if hasattr(exit_time, 'strftime'):
            exit_time = exit_time.strftime('%d/%m/%Y %H:%M')
        elif not exit_time:
            exit_time = '—'

        reason = log.get('exitReason', '—') or '—'
        if reason == 'Otro' and log.get('exitReasonDetail'):
            reason = f"Otro: {log['exitReasonDetail']}"

        status_class = 'status-entered' if log.get('status') == 'entered' else 'status-exited'
        status_text = 'Dentro' if log.get('status') == 'entered' else 'Salió'

        rows_html += f"""
        <tr>
            <td>{log.get('personnelName', '')}</td>
            <td>{log.get('personnelDni', '')}</td>
            <td>{log.get('shiftName', '')}</td>
            <td>{entry_time}</td>
            <td>{exit_time}</td>
            <td>{log.get('duration', '')}</td>
            <td>{reason}</td>
            <td>{log.get('areaName', '')}</td>
            <td class="{status_class}">{status_text}</td>
        </tr>
        """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: 'Helvetica', 'Arial', sans-serif; font-size: 9px; color: #333; }}
            .header {{ text-align: center; margin-bottom: 20px; border-bottom: 3px solid #1B3A5C; padding-bottom: 10px; }}
            .header h1 {{ color: #1B3A5C; font-size: 16px; margin: 0; }}
            .header p {{ color: #666; font-size: 10px; margin: 4px 0 0; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background-color: #1B3A5C; color: white; padding: 6px 4px; text-align: left; font-size: 8px; text-transform: uppercase; }}
            td {{ padding: 5px 4px; border-bottom: 1px solid #ddd; font-size: 8px; }}
            tr:nth-child(even) {{ background-color: #f8f9fa; }}
            .status-entered {{ color: #16a34a; font-weight: bold; }}
            .status-exited {{ color: #6b7280; }}
            .footer {{ margin-top: 20px; text-align: right; font-size: 8px; color: #999; border-top: 1px solid #ddd; padding-top: 8px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>REPORTE DE CONTROL DE ENTRADA Y SALIDA</h1>
            <p>SENATI — Sistema de Seguridad</p>
            <p>Período: {date_from or 'Inicio'} — {date_to or 'Actual'} | Total registros: {len(logs)}</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Guardia</th>
                    <th>DNI</th>
                    <th>Turno</th>
                    <th>Entrada</th>
                    <th>Salida</th>
                    <th>Duración</th>
                    <th>Motivo</th>
                    <th>Área</th>
                    <th>Estado</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        <div class="footer">
            Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Sistema de Control de Seguridad SENATI
        </div>
    </body>
    </html>
    """

    try:
        from xhtml2pdf import pisa
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"reporte_seguridad_{date_from or 'all'}_{date_to or 'all'}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Generate PDF directly to the response object
        pisa_status = pisa.CreatePDF(
            html_content, dest=response
        )

        if pisa_status.err:
            return HttpResponse('Hubo un error al generar el PDF.', status=500)
            
        return response

    except Exception as e:
        return HttpResponse(
            f"Error al generar PDF: {str(e)}",
            status=500,
        )
